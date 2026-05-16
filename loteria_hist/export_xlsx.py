"""Exportación de sorteos del periodo seleccionado a Excel (.xlsx)."""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from loteria_hist.repository import SorteoVista


# (columna, tipo_numero, prefijo_columna, cantidad)
_ESQUEMAS: dict[str, list[tuple[str, str, str, int]]] = {
    "euromillones": [
        ("principal", "principal", "N", 5),
        ("estrella", "estrella", "E", 2),
    ],
    "bonoloto": [
        ("principal", "principal", "N", 6),
        ("complementario", "complementario", "C", 1),
        ("reintegro", "reintegro", "R", 1),
    ],
    "primitiva": [
        ("principal", "principal", "N", 6),
        ("complementario", "complementario", "C", 1),
        ("reintegro", "reintegro", "R", 1),
    ],
}

_TITULOS_JUEGO = {
    "euromillones": "Euromillones",
    "bonoloto": "Bonoloto",
    "primitiva": "La Primitiva",
}


def parse_premio_eur(text: str | None) -> float | None:
    """Convierte textos tipo '17.000.000 €' a número."""
    if not text:
        return None
    s = re.sub(r"[^\d,.\-]", "", text.strip())
    if not s:
        return None
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def columnas_exportacion(juego: str) -> list[str]:
    cols = ["Fecha_Sorteo", "Dia_Semana"]
    for _tipo, _kind, pref, n in _ESQUEMAS.get(juego, []):
        cols.extend(f"{pref}{i}" for i in range(1, n + 1))
    cols.extend(["Premio_texto", "Premio_EUR"])
    return cols


def _fila_sorteo(juego: str, s: SorteoVista) -> dict[str, Any]:
    row: dict[str, Any] = {
        "Fecha_Sorteo": s.fecha,
        "Dia_Semana": s.dia_semana or "",
        "Premio_texto": s.premio_bote or "",
        "Premio_EUR": parse_premio_eur(s.premio_bote),
    }
    for tipo, _kind, pref, n in _ESQUEMAS.get(juego, []):
        vals = sorted(s.numeros.get(tipo, []))
        for i in range(1, n + 1):
            row[f"{pref}{i}"] = vals[i - 1] if i <= len(vals) else None
    return row


def sorteos_a_filas(juego: str, sorteos: list[SorteoVista]) -> list[dict[str, Any]]:
    ordenados = sorted(sorteos, key=lambda s: (s.fecha, s.id))
    return [_fila_sorteo(juego, s) for s in ordenados]


def _auto_ancho(ws, max_col: int, desde_fila: int = 1) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in ws.iter_rows(
            min_row=desde_fila, max_row=ws.max_row, min_col=col, max_col=col
        ):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, min(len(str(val)) + 2, 40))
        ws.column_dimensions[letter].width = max_len


def exportar_sorteos_periodo(
    path: str,
    juego: str,
    periodo_etiqueta: str,
    sorteos: list[SorteoVista],
) -> int:
    """
    Escribe un .xlsx con hoja de sorteos del periodo.
    Devuelve el número de filas exportadas.
    """
    filas = sorteos_a_filas(juego, sorteos)
    if not filas:
        raise ValueError("No hay sorteos en el periodo seleccionado")

    headers = columnas_exportacion(juego)
    titulo_juego = _TITULOS_JUEGO.get(juego, juego)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sorteos"[:31]

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=13, color="1E3A5F")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=f"{titulo_juego} — {periodo_etiqueta}")
    ws["A1"].font = title_font
    ws["A1"].alignment = left

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(
        row=2,
        column=1,
        value=f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · {len(filas)} sorteo(s)",
    )
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = left

    header_row = 4
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for r_idx, fila in enumerate(filas, start=header_row + 1):
        for c_idx, key in enumerate(headers, start=1):
            val = fila.get(key)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if key == "Fecha_Sorteo" and val:
                try:
                    d = date.fromisoformat(str(val))
                    cell.value = d
                    cell.number_format = "DD/MM/YYYY"
                except ValueError:
                    cell.value = val
                cell.alignment = center
            elif key.startswith(("N", "E", "C", "R")) and key[1:].isdigit():
                cell.alignment = center
                if val is not None:
                    cell.number_format = "0"
            elif key == "Premio_EUR" and val is not None:
                cell.number_format = '#,##0.00" €"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif key == "Premio_texto":
                cell.alignment = left
            else:
                cell.alignment = left

        if r_idx % 2 == 0:
            alt = PatternFill("solid", fgColor="F4F6F8")
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = alt

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"
    )

    _auto_ancho(ws, len(headers), desde_fila=header_row)

    wb.save(path)
    return len(filas)
